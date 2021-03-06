import discord
import asyncio
import random
import openpyxl
import requests
from captcha.image import ImageCaptcha
from bs4 import BeautifulSoup
from urllib.request import urlopen
from datetime import datetime

client = discord.Client()

@client.event
async def on_ready():
        print(client.user.id)
        print("ready")
        game = discord.Game("!도움말 ")
        await client.change_presence(status=discord.Status.online,activity=game)

@client.event
async def on_message(message):
    if message.content.startswith("!chatlog user"):
        if message.author.id == 570211549143105538:
            await message.channel.send("채팅로그불러오기를 시작합니다. 채팅로그를 불러오는 도중에 채팅을 치시면 에러 혹은 불러오기가 안될 수 있습니다.")
            author = message.content[14:32]
            file = openpyxl.load_workbook("채팅로그.xlsx")
            sheet = file.active
            msg = "채팅로그를 불러온 결과입니다."
            foot = datetime.today().strftime("%Y년 %m월 %d일 %H시 %M분 %S초")
            embed = discord.Embed(title="채팅로그", description=msg,color=0xDF013A)  # Embed의 기본 틀(색상, 메인 제목, 설명)을 잡아줍니다
            embed.set_footer(text=foot)  # 하단에 들어가는 조그마한 설명을 잡아줍니다
            i = 1
            while True:
                if sheet["A" + str(i)].value == str(author):
                    if sheet["E" + str(i)].value == str(message.guild.id):
                        if sheet["D" + str(i)].value == str(message.channel.id):
                            embed.add_field(name=sheet["B" + str(i)].value + "(" + sheet["A" + str(i)].value + ")", value=sheet["C" + str(i)].value, inline=True)

                if sheet["A" + str(i)].value == None:
                     break
                i += 1
            await message.channel.send(embed=embed)

        if message.author.id == 345265069132742657:
            await message.channel.send("채팅로그불러오기를 시작합니다. 채팅로그를 불러오는 도중에 채팅을 치시면 에러 혹은 불러오기가 안될 수 있습니다.")
            author = message.content[14:32]
            file = openpyxl.load_workbook("채팅로그.xlsx")
            sheet = file.active
            msg = "채팅로그를 불러온 결과입니다."
            foot = datetime.today().strftime("%Y년 %m월 %d일 %H시 %M분 %S초")
            embed = discord.Embed(title="채팅로그", description=msg, color=0xDF013A)  # Embed의 기본 틀(색상, 메인 제목, 설명)을 잡아줍니다
            embed.set_footer(text=foot)  # 하단에 들어가는 조그마한 설명을 잡아줍니다
            i = 1
            while True:
                if sheet["A" + str(i)].value == str(author):
                    if sheet["E" + str(i)].value == str(message.guild.id):
                        if sheet["D" + str(i)].value == str(message.channel.id):
                            embed.add_field(name=sheet["B" + str(i)].value + "(" + sheet["A" + str(i)].value + ")",
                                            value=sheet["C" + str(i)].value, inline=True)

                if sheet["A" + str(i)].value == None:
                    break
                i += 1
            await message.channel.send(embed=embed)

    if message.content.startswith("!chatlog name"):
        if message.author.id == 570211549143105538:
            await message.channel.send("채팅로그불러오기를 시작합니다. 채팅로그를 불러오는 도중에 채팅을 치시면 에러 혹은 불러오기가 안될 수 있습니다.")
            author = message.content[14:]
            file = openpyxl.load_workbook("채팅로그.xlsx")
            sheet = file.active
            msg = "채팅로그를 불러온 결과입니다."
            foot = datetime.today().strftime("%Y년 %m월 %d일 %H시 %M분 %S초")
            embed = discord.Embed(title="채팅로그", description=msg, color=0xDF013A)  # Embed의 기본 틀(색상, 메인 제목, 설명)을 잡아줍니다
            embed.set_footer(text=foot)  # 하단에 들어가는 조그마한 설명을 잡아줍니다
            i = 1
            while True:
                if sheet["B" + str(i)].value == str(author):
                    if sheet["E" + str(i)].value == str(message.guild.id):
                        if sheet["D" + str(i)].value == str(message.channel.id):
                            embed.add_field(name=sheet["B" + str(i)].value + "(" + sheet["A" + str(i)].value + ")",
                                            value=sheet["C" + str(i)].value, inline=True)

                if sheet["A" + str(i)].value == None:
                    break
                i += 1
            await message.channel.send(embed=embed)
    if message.author.id == 345265069132742657:
        await message.channel.send("채팅로그불러오기를 시작합니다. 채팅로그를 불러오는 도중에 채팅을 치시면 에러 혹은 불러오기가 안될 수 있습니다.")
        author = message.content[14:]
        file = openpyxl.load_workbook("채팅로그.xlsx")
        sheet = file.active
        msg = "채팅로그를 불러온 결과입니다."
        foot = datetime.today().strftime("%Y년 %m월 %d일 %H시 %M분 %S초")
        embed = discord.Embed(title="채팅로그", description=msg, color=0xDF013A)  # Embed의 기본 틀(색상, 메인 제목, 설명)을 잡아줍니다
        embed.set_footer(text=foot)  # 하단에 들어가는 조그마한 설명을 잡아줍니다
        i = 1
        while True:
            if sheet["B" + str(i)].value == str(author):
                if sheet["E" + str(i)].value == str(message.guild.id):
                    if sheet["D" + str(i)].value == str(message.channel.id):
                        embed.add_field(name=sheet["B" + str(i)].value + "(" + sheet["A" + str(i)].value + ")",
                                        value=sheet["C" + str(i)].value, inline=True)

            if sheet["A" + str(i)].value == None:
                break
            i += 1
        await message.channel.send(embed=embed)

    if message.content.startswith("!chatlog chat"):
        if message.author.id == 570211549143105538:
            await message.channel.send("채팅로그불러오기를 시작합니다. 채팅로그를 불러오는 도중에 채팅을 치시면 에러 혹은 불러오기가 안될 수 있습니다.")
            author = message.content[14:]
            file = openpyxl.load_workbook("채팅로그.xlsx")
            sheet = file.active
            msg = "채팅로그를 불러온 결과입니다."
            foot = datetime.today().strftime("%Y년 %m월 %d일 %H시 %M분 %S초")
            embed = discord.Embed(title="채팅로그", description=msg, color=0xDF013A)  # Embed의 기본 틀(색상, 메인 제목, 설명)을 잡아줍니다
            embed.set_footer(text=foot)  # 하단에 들어가는 조그마한 설명을 잡아줍니다
            i = 1
            while True:
                if sheet["C" + str(i)].value == str(author):
                    if sheet["E" + str(i)].value == str(message.guild.id):
                        if sheet["D" + str(i)].value == str(message.channel.id):
                            embed.add_field(name=sheet["B" + str(i)].value + "(" + sheet["A" + str(i)].value + ")",
                                            value=sheet["C" + str(i)].value, inline=True)

                if sheet["A" + str(i)].value == None:
                    break
                i += 1
            await message.channel.send(embed=embed)
    if message.author.id == 345265069132742657:
        await message.channel.send("채팅로그불러오기를 시작합니다. 채팅로그를 불러오는 도중에 채팅을 치시면 에러 혹은 불러오기가 안될 수 있습니다.")
        author = message.content[14:]
        file = openpyxl.load_workbook("채팅로그.xlsx")
        sheet = file.active
        msg = "채팅로그를 불러온 결과입니다."
        foot = datetime.today().strftime("%Y년 %m월 %d일 %H시 %M분 %S초")
        embed = discord.Embed(title="채팅로그", description=msg, color=0xDF013A)  # Embed의 기본 틀(색상, 메인 제목, 설명)을 잡아줍니다
        embed.set_footer(text=foot)  # 하단에 들어가는 조그마한 설명을 잡아줍니다
        i = 1
        while True:
            if sheet["C" + str(i)].value == str(author):
                if sheet["E" + str(i)].value == str(message.guild.id):
                    if sheet["D" + str(i)].value == str(message.channel.id):
                        embed.add_field(name=sheet["B" + str(i)].value + "(" + sheet["A" + str(i)].value + ")",value=sheet["C" + str(i)].value, inline=True)

            if sheet["A" + str(i)].value == None:
                break
            i += 1
        await message.channel.send(embed=embed)

    if message.content == "!코로나현황":
        response = requests.get('https://search.naver.com/search.naver?sm=tab_hty.top&where=nexearch&query=코로나')
        readerhtml = response.text
        soup = BeautifulSoup(readerhtml, 'lxml')
        data1 = soup.find('div', class_='graph_view')
        data2 = data1.findAll('div', class_='box')
        data3 = data1.findAll('div', class_='box bottom')
        checked = data2[0].find('p', class_='txt').find('strong', class_='num').text
        checking = data2[2].find('p', class_='txt').find('strong', class_='num').text
        free = data3[0].find('p', class_='txt').find('strong', class_='num').text
        die = data3[1].find('p', class_='txt').find('strong', class_='num').text
        wasup = soup.find('div', class_='csp_notice_info').find('p').find_all(text=True, recursive=True)
        # ===============================
        coembed = discord.Embed(color=0x192131, title='코로나현황', description=f'{wasup[1]}')
        coembed.add_field(name="확진자", value=f'{checked}명', inline=True)
        coembed.add_field(name="격리해제", value=f'{free}명', inline=True)
        coembed.add_field(name="검사중", value=f'{checking}명', inline=True)
        coembed.add_field(name="사망자", value=f'{die}명', inline=True)
        coembed.set_footer(text=datetime.today().strftime("%Y년 %m월 %d일 %H:%M:%S | https://github.com/sus5373/woozzange-bot-sows"))
        await message.channel.send(embed=coembed)

    if message.content == ("!도움말 2"):
        msg = "도움말 2 페이지 입니다."
        foot = datetime.today().strftime("%Y년 %m월 %d일 %H시 %M분 %S초")
        embed = discord.Embed(title="도움말", description=msg, color=0xFAED7D)  # Embed의 기본 틀(색상, 메인 제목, 설명)을 잡아줍니다
        embed.set_footer(text=foot)  # 하단에 들어가는 조그마한 설명을 잡아줍니다
        embed.add_field(name="!강아지", value="랜덤으로 강아지 사진이 나옵니다.", inline=True)
        embed.add_field(name="!고양이", value="랜덤으로 고양이 사진이 나옵니다.", inline=True)
        embed.add_field(name="!네코", value="랜덤으로 에니메이션 주인공이 나옵니다.", inline=True)
        await message.channel.send(embed=embed)
    
    if message.content.startswith('!고양이'):
        embed = discord.Embed(
            title='고양이는',
            description='멍멍',
            colour=discord.Colour.green()
        )

        urlBase = 'https://loremflickr.com/320/240?lock='
        randomNum = random.randrange(1, 30977)
        urlF = urlBase + str(randomNum)
        embed.set_image(url=urlF)
        await message.channel.send(embed=embed)

    if message.content.startswith('!강아지'):
        embed = discord.Embed(
            title='강아지는',
            description='야옹야옹',
            colour=discord.Colour.green()
        )

        urlBase = 'https://loremflickr.com/320/240/dog?lock='
        randomNum = random.randrange(1, 30977)
        urlF = urlBase + str(randomNum)
        embed.set_image(url=urlF)
        await message.channel.send(embed=embed)

    if message.content.startswith('!네코'):
        embed = discord.Embed(
            colour=discord.Colour.green()
        )
        embed2 = discord.Embed(
            colour=discord.Colour.green()
        )
        embed3 = discord.Embed(
            colour=discord.Colour.green()
        )
        randomnumber = random.randrange(100, 407)
        randomgiho = random.randrange(1, 3)
        print('?번째사진 : ' + str(randomnumber))
        print('기호 : ' + str(randomgiho))
        strandomnumber = str(randomnumber)
        file1 = '.png'
        file2 = '.jpg'
        file3 = '.jpeg'
        giho = '_'
        if randomgiho == 1:
            urlbase1 = "https://cdn.nekos.life/neko/neko" + strandomnumber + file1
            urlbase2 = "https://cdn.nekos.life/neko/neko" + strandomnumber + file2
            urlbase3 = "https://cdn.nekos.life/neko/neko" + strandomnumber + file3
            embed.set_image(url=urlbase1)
            embed2.set_image(url=urlbase2)
            embed3.set_image(url=urlbase3)
            await message.channel.send(embed=embed)
            await message.channel.send(embed=embed2)
            await message.channel.send(embed=embed3)
        else:
            urlbase_1 = "https://cdn.nekos.life/neko/neko" + giho + strandomnumber + file1
            urlbase_2 = "https://cdn.nekos.life/neko/neko" + giho + strandomnumber + file2
            urlbase_3 = "https://cdn.nekos.life/neko/neko" + giho + strandomnumber + file3
            embed.set_image(url=urlbase_1)
            embed2.set_image(url=urlbase_2)
            embed3.set_image(url=urlbase_3)
            await message.channel.send(embed=embed)
            await message.channel.send(embed=embed2)
            await message.channel.send(embed=embed3)

    if message.content == "!핑":
        latancy = client.latency
        await message.channel.send(f'퐁! {round(latancy*1000)}ms')

    

    if message.content == ("!도움말 1"):
        msg = "도움말 1 페이지 입니다."
        foot = datetime.today().strftime("%Y년 %m월 %d일 %H시 %M분 %S초")
        embed = discord.Embed(title="도움말", description=msg, color=0xFAED7D)  # Embed의 기본 틀(색상, 메인 제목, 설명)을 잡아줍니다
        embed.set_footer(text=foot)  # 하단에 들어가는 조그마한 설명을 잡아줍니다
        embed.add_field(name="!핑", value="봇의 핑을 알려드립니다.", inline=True)
        embed.add_field(name="!코로나현황", value="현재 대한민국의 코로나 현황을 알려드립니다.", inline=True)
        embed.add_field(name="!도움말", value="봇의 명령어들을 알려드립니다.", inline=True)
        embed.add_field(name="!날자", value="현재 날자를 알려드립니다.", inline=True)
        embed.add_field(name="!id", value="자신의 아이디, 채널아이디, 서버아이디를 알려드립니다.", inline=True)
        embed.add_field(name="!dm", value="해당 유저에게 dm을 보냅니다.", inline=True)
        embed.add_field(name="!인증", value="캡챠(인증)를 합니다.", inline=True)
        embed.add_field(name="!file", value="우섭님의 컴퓨터에 있는 파일을 불러옵니다.", inline=True)
        embed.add_field(name="!봇말해", value="봇이 내용을 말합니다.", inline=True)
        await message.channel.send(embed=embed)

    if message.content.startswith("!날자"):
        await message.channel.send(datetime.today().strftime("%Y년 %m월 %d일 %H시 %M분 %S초"))

    if message.content.startswith("!공지"):
        file = openpyxl.load_workbook("서버목록.xlsx")
        sheet = file.active
        i = 1
        while True:
            if sheet["A" + str(i)].value == str(message.guild.id):
                channel = sheet["B" + str(i)].value
                msg = message.content[4:]
                foot = datetime.today().strftime("%Y년 %m월 %d일 %H시 %M분 %S초")
                embed = discord.Embed(title="우짱이봇 공지", description=msg, color=0xFFFF00)  # Embed의 기본 틀(색상, 메인 제목, 설명)을 잡아줍니다
                embed.set_footer(text=foot)  # 하단에 들어가는 조그마한 설명을 잡아줍니다
                await client.get_channel(int(channel)).send(embed=embed)
                break
            if sheet["A" + str(i)].value == None:
                await message.channel.send("공지채널이 설정되어 있지 않습니다.")
                break
            i += 1

    if message.content.startswith("!설정공지"):
        channel = int(message.content[6:26])
        file = openpyxl.load_workbook("서버목록.xlsx")
        sheet = file.active
        i = 1
        while True:
            if sheet["A" + str(i)].value == str(message.guild.id):
                sheet["B" + str(i)].value = str(channel)
                file.save("서버목록.xlsx")
                await message.channel.send("정상적으로 설정되었습니다!")
                break
            if sheet["A" + str(i)].value == None:
                sheet["A" + str(i)].value = str(message.guild.id)
                sheet["B" + str(i)].value = str(channel)
                file.save("서버목록.xlsx")
                await message.channel.send("정상적으로 설정되었습니다!")
                break
            i += 1

    if message.content == ("!도움말"):
        msg = "'!도움말 + 쪽수' 를 입력해 주세요."
        foot = datetime.today().strftime("%Y년 %m월 %d일 %H시 %M분 %S초")
        embed = discord.Embed(title="도움말", description=msg, color=0xFAED7D)  # Embed의 기본 틀(색상, 메인 제목, 설명)을 잡아줍니다
        embed.set_footer(text=foot)  # 하단에 들어가는 조그마한 설명을 잡아줍니다
        await message.channel.send(embed=embed)

    if message.content.startswith("!id"):
        msg = "내아이디 : " + str(message.author.id) + "\n채널아이디 :" + str(message.channel.id) + "\n서버아이디 :" + str(message.guild.id)
        foot = datetime.today().strftime("%Y년 %m월 %d일 %H시 %M분 %S초")
        embed = discord.Embed(title="ID", description=msg, color=0x65FF5E)  # Embed의 기본 틀(색상, 메인 제목, 설명)을 잡아줍니다
        embed.set_footer(text=foot)  # 하단에 들어가는 조그마한 설명을 잡아줍니다
        await message.channel.send(embed=embed)  # embed를 포함 한 채로 메시지를 전송합니다.

    if message.content.startswith("!dm"):
        author = message.guild.get_member(int(message.content[7:25]))
        msg = message.content[27:]
        await author.send(msg)

    if message.content.startswith("!인증"):
        Image_chaptcha = ImageCaptcha()
        a = ""
        for i in range(5):
            a += str(random.randint(0,9))
        name = str(message.author.id) + ".png"
        Image_chaptcha.write(a, name)
        await message.channel.send(file=discord.File(name))
        def check(msg):
            return msg.author == message.author and msg.channel == message.channel

        try:
            msg = await client.wait_for("message",timeout=10, check=check)
        except:
            await message.channel.send("시간초과입니다")
            return

        if msg.content == a:
            await message.channel.send("정답입니다.")
        else:
            await message.channel.send("오답입니다.")


    if message.content.startswith("!file"):
        pic = message.content[6:]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith(""):
        file = openpyxl.load_workbook("채팅로그.xlsx")
        sheet = file.active
        i = 1
        while True:
            if sheet["A" + str(i)].value == None:
                sheet["A" + str(i)].value = str(message.author.id)
                sheet["B" + str(i)].value = str(message.author.name)
                sheet["C" + str(i)].value = str(message.content)
                sheet["D" + str(i)].value = str(message.channel.id)
                sheet["E" + str(i)].value = str(message.guild.id)
                file.save("채팅로그.xlsx")
                break
            i += 1

    if message.content.startswith("!정지추가"):
        file = openpyxl.load_workbook("관리자.xlsx")
        sheet = file.active
        i = 1
        while True:
            if sheet["A" + str(i)].value == str(message.author.id):
                author = message.content[6:24]
                file = openpyxl.load_workbook("정지목록.xlsx")
                sheet = file.active
                j = 1
                while True:
                    if sheet["A" + str(j)].value == None:
                        sheet["A" + str(i)].value = str(author)
                        await message.channel.send("정상적으로 추가했습니다.")
                        file.save("정지목록.xlsx")
                        break
                    i += 1
                break
            i+=1

    if message.content.startswith("!봇말해"):
        msg = message.content[5:]
        await message.channel.send(msg)


client.run("TOKEN")
